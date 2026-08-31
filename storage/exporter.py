"""
Générateur et exportateur de fichiers Excel professionnels (openpyxl).
Produit le fichier 'contacts_stage.xlsx' avec mise en forme, colorations de statut,
tri intelligent par entreprise/priorité et déduplication stricte.
"""

from pathlib import Path
from typing import Any, Dict, List, Optional
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from config import config
from core.monitoring.audit_logger import audit_logger
from enricher.lead_scorer import lead_scorer
from storage.db_manager import db_manager


class ExcelExporter:
    def __init__(self, output_path: Optional[Path] = None):
        self.output_path = output_path or config.OUTPUT_EXCEL_PATH
        self.history_path = self.output_path.parent / "contacts_historique.xlsx"

    def archive_current_session_file(self) -> None:
        """
        Déplace l'actuel contacts_stage.xlsx vers contacts_historique.xlsx
        au démarrage d'une nouvelle session, fusionne avec l'historique existant,
        et réinitialise le fichier de session pour la nouvelle recherche.
        """
        import shutil
        data_dir = self.output_path.parent
        data_dir.mkdir(parents=True, exist_ok=True)
        main_path = self.output_path
        history_path = self.history_path

        if main_path.exists() and main_path.stat().st_size > 1000:
            try:
                # Si un historique existe déjà, on sauvegarde la base globale cumulée
                all_leads = db_manager.get_all_leads()
                if all_leads:
                    self.export_leads(all_leads, destination=history_path)
                else:
                    shutil.copy2(str(main_path), str(history_path))
                
                audit_logger.log_event("SESSION_ARCHIVE", f"Session précédente archivée dans {history_path.name}")
                
                # Réinitialisation du fichier de session actuelle
                self.export_leads([], destination=main_path)
            except Exception as e:
                audit_logger.log_event("SESSION_ARCHIVE_ERR", f"Erreur archivage : {e}")

    def export_leads(self, leads: List[Dict[str, Any]], destination: Optional[Path] = None) -> Path:
        """
        Génère ou met à jour immédiatement le fichier Excel avec une liste de leads (leads de session ou base complète).
        Écrit sur disque de manière synchronisée et résistante aux coupures.
        """
        dest_path = destination or self.output_path
        dest_path.parent.mkdir(parents=True, exist_ok=True)

        # 1. Déduplication en mémoire stricte sur (Nom, Prénom, Entreprise)
        seen = set()
        clean_leads: List[Dict[str, Any]] = []
        for lead in leads:
            fn = lead.get("first_name", "").strip().lower()
            ln = lead.get("last_name", "").strip().lower()
            comp = lead.get("company", "").strip().lower()
            key = (fn, ln, comp)
            if key not in seen and fn and ln:
                seen.add(key)
                score, etoiles, _ = lead_scorer.calculate_score(lead)
                lead["priority_stars"] = etoiles
                lead["relevance_score"] = score
                clean_leads.append(lead)

        # 2. Tri intelligent : d'abord par Entreprise, puis par Score de Pertinence décroissant
        clean_leads.sort(key=lambda x: (x.get("company", "").upper(), -x.get("relevance_score", 0), -x.get("confidence_score", 0)))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contacts Stage"

        # Définition des en-têtes
        headers = [
            "Prénom",
            "Nom",
            "Poste",
            "Entreprise",
            "Priorité",
            "Email (proposé)",
            "Email (alternatif 1)",
            "Email (alternatif 2)",
            "Score de confiance",
            "Statut",
            "MX vérifié",
            "URL LinkedIn",
            "Mots-clés matchés"
        ]

        # Styles
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        fill_valide = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")     # Vert clair
        fill_non_verif = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Rouge/Orange clair
        fill_a_verifier = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Jaune clair

        # Écriture de la ligne d'en-tête
        ws.append(headers)
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        # Écriture des données nettoyées et triées
        for row_idx, lead in enumerate(clean_leads, start=2):
            score_str = f"{lead.get('confidence_score', 0)}%"
            row_values = [
                lead.get("first_name", ""),
                lead.get("last_name", ""),
                lead.get("job_title", ""),
                lead.get("company", ""),
                lead.get("priority_stars", "★☆☆"),
                lead.get("proposed_email", ""),
                lead.get("alt_email_1", ""),
                lead.get("alt_email_2", ""),
                score_str,
                lead.get("status", "À vérifier"),
                lead.get("mx_verified", "Non"),
                lead.get("profile_url", ""),
                lead.get("matched_keywords", "")
            ]
            ws.append(row_values)

            # Application des bordures et colorations conditionnelles
            status_val = str(lead.get("status", ""))
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.border = thin_border
                cell.alignment = center_align if col_num in (5, 9, 10, 11) else left_align

                if col_num == 10:  # Colonne Statut
                    if status_val.startswith("Validé"):
                        cell.fill = fill_valide
                    elif "Non vérifiable" in status_val or "Invalide" in status_val:
                        cell.fill = fill_non_verif
                    elif "À vérifier" in status_val:
                        cell.fill = fill_a_verifier

        # Ajustement automatique de la largeur des colonnes
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Figer la ligne supérieure et activer les filtres automatiques
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        wb.save(str(dest_path))
        audit_logger.log_event("EXCEL_EXPORT", f"Exportation de {len(clean_leads)} leads nettoyés dans {dest_path}")
        return dest_path

    def export_from_db(self, destination: Optional[Path] = None, db: Optional[Any] = None) -> Path:
        """Exporte tous les leads de la base de données vers le fichier Excel."""
        target_db = db or db_manager
        raw_leads = target_db.get_all_leads()
        return self.export_leads(raw_leads, destination=destination)

    @staticmethod
    def list_export_history() -> List[Dict[str, Any]]:
        """
        Retourne l'archive globale et de la session précédente avec leurs métadonnées.
        """
        from datetime import datetime
        results = []
        data_dir = config.DATA_DIR
        
        hist_path = data_dir / "contacts_historique.xlsx"
        if hist_path.exists() and hist_path.stat().st_size > 0:
            try:
                stat = hist_path.stat()
                dt = datetime.fromtimestamp(stat.st_mtime).strftime("%d/%m/%Y %H:%M")
                results.append({
                    "filename": "contacts_historique.xlsx (Base Cumulée)",
                    "filepath": str(hist_path),
                    "size_kb": round(stat.st_size / 1024, 1),
                    "date": dt
                })
            except Exception:
                pass

        archive_path = config.EXPORTS_DIR / "contacts_session_precedente.xlsx"
        if archive_path.exists() and archive_path.stat().st_size > 0:
            try:
                stat = archive_path.stat()
                dt = datetime.fromtimestamp(stat.st_mtime).strftime("%d/%m/%Y %H:%M")
                results.append({
                    "filename": "contacts_session_precedente.xlsx",
                    "filepath": str(archive_path),
                    "size_kb": round(stat.st_size / 1024, 1),
                    "date": dt
                })
            except Exception:
                pass

        return results


excel_exporter = ExcelExporter()
