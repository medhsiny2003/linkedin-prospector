"""
Tests unitaires pour l'exportateur Excel (openpyxl).
"""

import os
import tempfile
from pathlib import Path
import openpyxl
import pytest
from storage.db_manager import DBManager
from storage.exporter import ExcelExporter


@pytest.fixture
def temp_environment():
    fd_db, db_path_str = tempfile.mkstemp(suffix=".db")
    os.close(fd_db)
    db_path = Path(db_path_str)

    fd_xl, xl_path_str = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd_xl)
    xl_path = Path(xl_path_str)

    db = DBManager(db_path=db_path)
    exporter = ExcelExporter(output_path=xl_path)

    yield db, exporter, xl_path

    try:
        if db_path.exists():
            os.remove(db_path)
    except Exception:
        pass

    try:
        if xl_path.exists():
            os.remove(xl_path)
    except Exception:
        pass


def test_excel_export_structure(temp_environment):
    db, exporter, xl_path = temp_environment

    lead = {
        "first_name": "Claire",
        "last_name": "Lemoine",
        "job_title": "Talent Acquisition",
        "company": "Safran",
        "profile_url": "https://www.linkedin.com/in/claire-lemoine",
        "proposed_email": "claire.lemoine@safran-group.com",
        "alt_email_1": "clairelemoine@safran-group.com",
        "alt_email_2": "c.lemoine@safran-group.com",
        "confidence_score": 95,
        "status": "Validé",
        "mx_verified": "Oui",
        "matched_keywords": "Safran, RH"
    }
    db.save_lead(lead)

    # Exporter avec l'instance db de test
    res_path = exporter.export_from_db(destination=xl_path, db=db)
    assert res_path.exists()

    # Vérification avec openpyxl
    wb = openpyxl.load_workbook(str(res_path))
    ws = wb.active
    assert ws.title == "Contacts Stage"

    # Vérification de l'en-tête
    headers = [cell.value for cell in ws[1]]
    assert "Prénom" in headers
    assert "Email (proposé)" in headers
    assert "Score de confiance" in headers
    assert "MX vérifié" in headers

    # Vérification des données ligne 2
    row2 = [cell.value for cell in ws[2]]
    assert "Claire" in row2
    assert "Safran" in row2
    assert "claire.lemoine@safran-group.com" in row2
    assert "95%" in row2
