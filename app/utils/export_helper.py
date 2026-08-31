"""
Générateur de flux de fichiers Excel et CSV en mémoire pour st.download_button.
"""

import io
from typing import List, Optional
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd


def generate_excel_bytes(df: pd.DataFrame, selected_columns: Optional[List[str]] = None) -> bytes:
    """Génère un fichier Excel (.xlsx) stylisé en mémoire (BytesIO)."""
    if selected_columns:
        valid_cols = [c for c in selected_columns if c in df.columns]
        export_df = df[valid_cols]
    else:
        export_df = df.copy()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # Styles LinkedIn
    header_fill = PatternFill(start_color="0A66C2", end_color="0A66C2", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="E0E4E8"),
        right=Side(style="thin", color="E0E4E8"),
        top=Side(style="thin", color="E0E4E8"),
        bottom=Side(style="thin", color="E0E4E8")
    )

    fill_valide = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_non_verif = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_a_verifier = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    headers = list(export_df.columns)
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    status_col_idx = headers.index("status") + 1 if "status" in headers else None

    for row_idx, row_data in enumerate(export_df.itertuples(index=False), start=2):
        ws.append(list(row_data))
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.border = thin_border
            cell.alignment = left_align

            if status_col_idx and col_num == status_col_idx:
                cell.alignment = center_align
                val = str(cell.value or "")
                if "Validé" in val:
                    cell.fill = fill_valide
                elif "Non vérifiable" in val:
                    cell.fill = fill_non_verif
                elif "À vérifier" in val:
                    cell.fill = fill_a_verifier

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_csv_bytes(df: pd.DataFrame, selected_columns: Optional[List[str]] = None) -> bytes:
    """Génère un flux CSV UTF-8 avec BOM (compatible Excel français avec séparateur point-virgule)."""
    if selected_columns:
        valid_cols = [c for c in selected_columns if c in df.columns]
        export_df = df[valid_cols]
    else:
        export_df = df.copy()

    csv_str = export_df.to_csv(index=False, sep=";", encoding="utf-8-sig")
    return csv_str.encode("utf-8-sig")


def get_export_history() -> List[dict]:
    """Retourne l'archive unique de la session précédente."""
    from storage.exporter import excel_exporter
    return excel_exporter.list_export_history()

